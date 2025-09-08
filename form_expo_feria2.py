# -*- coding: utf-8 -*-
# FORMULARIO DATOS EXPO FERIA - versión integral y robusta
# MECANICO / DISTRIBUIDOR / CONSUMIDOR / PUNTAJE / PREMIOS / CONSULTA / STAND

import os, re, time
from pathlib import Path
import streamlit as st
from openpyxl import Workbook, load_workbook

st.set_page_config(page_title="Formulario Expo Feria", page_icon="📝", layout="centered")

# ===== RUTA DEL EXCEL (lista para nube) =====
# Usa variable de entorno EXCEL_DIR (ej. /data en Render/Railway).
# Por defecto guarda en la carpeta relativa 'data/' del proyecto.
EXCEL_DIR = os.environ.get("EXCEL_DIR", "data")
EXCEL_FILE = os.environ.get("EXCEL_FILE", "FORMULARIO DATOS EXPO FERIA.xlsx")
EXCEL_PATH = Path(EXCEL_DIR) / EXCEL_FILE
EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)

# ===== ENCABEZADOS =====
HEADERS = {
    "MECANICO": [
        "CODIGO","NOMBRE Y APELLIDO","RUC O CEDULA","TELEFONO","CORREO",
        "PROVINCIA","CANTON/CIUDAD","PARROQUIA","DIRECCION","REDES SOCIALES",
        "A QUE TE DEDICAS","EDAD","NOMBRE DE LA MECÁNICA","MECANICA Y LOCAL",
        "QUISIERAS QUE TE VISITEMOS","PRODUCTOS DE INTERES","STAND"
    ],
    "DISTRIBUIDOR": [
        "CODIGO","NOMBRE Y APELLIDO","CEDULA O RUC","TELEFONO","EDAD",
        "PROVINCIA","CANTON/CIUDAD","PARROQUIA","DIRECCION","CORREO",
        "REDES SOCIALES","A QUE TE DEDICAS","QUE REPUESTOS QUIERES DISTRIBUIR","STAND"
    ],
    "CONSUMIDOR": [
        "CODIGO","NOMBRE Y APELLIDO","CEDULA O RUC","TELEFONO","EDAD",
        "HOMBRE O MUJER","PROVINCIA","PROVINCIA","CANTON/CIUDAD","PARROQUIA",
        "DIRECCION","A QUE TE DEDICAS","MODELO DE MOTO QUE USAS",
        "QUE REPUESTO BUSCAS?","SI HAS COMPRANDO PRODCUTOS EXTREMEMAX?","STAND"
    ],
    "PROVINCIA": ["PROVINCIA","CANTON/CIUDAD","PARROQUIA"],
    "REGISTRO DE CODIGOS": ["CODIGO","PUNTAJE","RUC O CEDULA","NOMBRE","TELEFONO","TIPO","STAND"],
    "REGISTRO DE PREMIOS": ["CODIGO","PREMIO","RUC O CEDULA","NOMBRE","TELEFONO","TIPO","STAND"],
}

# ===== UTILIDADES EXCEL (robusto) =====
def _mtime(p: Path) -> float:
    try: return os.path.getmtime(p)
    except: return 0.0

def safe_load_workbook(path, read_only=False, data_only=False, tries=10, wait=0.4):
    last = None
    for _ in range(tries):
        try:
            return load_workbook(path, read_only=read_only, data_only=data_only)
        except PermissionError as e:
            last = e
            time.sleep(wait)
    raise last if last else PermissionError("No se pudo abrir el Excel (bloqueado).")

def safe_save_workbook(wb, path: Path, tries=30, wait=0.5):
    """
    1) Intenta guardar directo (Windows-friendly).
    2) Reintenta varias veces si está bloqueado.
    3) Si no se puede, guarda una COPIA con timestamp en la misma carpeta.
    """
    last = None
    for _ in range(tries):
        try:
            wb.save(path)
            try: wb.close()
            except: pass
            return
        except PermissionError as e:
            last = e; time.sleep(wait)
        except Exception as e:
            last = e; break

    # Fallback: copia con timestamp
    stamp = time.strftime("%Y%m%d-%H%M%S")
    alt = path.with_name(f"{path.stem}_copia_{stamp}{path.suffix}")
    try:
        wb.save(alt)
        try: wb.close()
        except: pass
        st.warning(f"⚠️ El archivo principal está bloqueado. Guardé una COPIA: {alt}")
        return
    except Exception as e2:
        try: wb.close()
        except: pass
        raise last or e2

def _norm_text(x):
    if x is None: return ""
    s=str(x).strip().upper()
    s=(s.replace("Á","A").replace("É","E").replace("Í","I")
         .replace("Ó","O").replace("Ú","U").replace("Ü","U").replace("Ñ","N"))
    return re.sub(r"\s+"," ", s)

def header_map(ws):
    head=next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {_norm_text(v):i for i,v in enumerate(head)}

def _sync_headers(ws, desired_headers):
    if ws.max_row==1 and ws["A1"].value is None:
        ws.append(desired_headers); return True
    head = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    existing = [str(h or "").strip() for h in head]
    existing_norm = [_norm_text(h) for h in existing]
    changed = False
    for h in desired_headers:
        if _norm_text(h) not in existing_norm:
            ws.cell(row=1, column=len(existing)+1, value=h)
            existing.append(h); existing_norm.append(_norm_text(h))
            changed = True
    return changed

def ensure_workbook(path: Path):
    if not path.exists():
        wb=Workbook()
        if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
        for name, headers in HEADERS.items():
            ws=wb.create_sheet(name); ws.append(headers)
        safe_save_workbook(wb, path); return
    try: wb=safe_load_workbook(path)
    except PermissionError: return
    changed=False
    for name, headers in HEADERS.items():
        if name not in wb.sheetnames:
            ws=wb.create_sheet(name); ws.append(headers); changed=True
        else:
            ws=wb[name]
            if _sync_headers(ws, headers): changed=True
    if changed: safe_save_workbook(wb, path)

def append_row(sheet: str, values: list) -> bool:
    ensure_workbook(EXCEL_PATH)
    try:
        wb = safe_load_workbook(EXCEL_PATH)
    except PermissionError:
        st.error("🔒 Cierra el Excel o pausa OneDrive e intenta de nuevo.")
        return False
    try:
        ws = wb[sheet]; ws.append(values)
        safe_save_workbook(wb, EXCEL_PATH)
        return True
    except PermissionError:
        st.error("🔒 No se pudo guardar (archivo bloqueado)."); return False
    except Exception as e:
        st.error(f"❗ Error inesperado al guardar: {e}"); return False

# ===== NORMALIZACIÓN / VALIDACIÓN =====
def norm_id(s):   return re.sub(r"\D+","", s or "")
def norm_phone(s):
    s = re.sub(r"\D+","", s or ""); return s[-10:] if len(s)>=10 else s
def norm_email(s): return (s or "").strip().lower()

def validar_cedula_ec(num):
    s=norm_id(num)
    if len(s)!=10: return False
    try:
        if not (1<=int(s[:2])<=24): return False
        if int(s[2])>=6: return False
        coef=[2,1,2,1,2,1,2,1,2]; tot=0
        for i in range(9):
            x=int(s[i])*coef[i]; tot+= x if x<10 else x-9
        dv=(10-(tot%10))%10
        return dv==int(s[9])
    except: return False

def validar_ruc_natural_ec(num):
    s=norm_id(num)
    return len(s)==13 and s.endswith("001") and validar_cedula_ec(s[:10])

def email_valido(e):
    if not e: return True
    return re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", e.strip()) is not None

# ===== ÍNDICE PROVINCIA → CANTÓN → PARROQUIA =====
@st.cache_data(show_spinner=False)
def load_province_index(path_str: str, mtime: float):
    index={}
    try:
        wb=safe_load_workbook(Path(path_str), read_only=True, data_only=True)
        if "PROVINCIA" not in wb.sheetnames: return {}
        ws=wb["PROVINCIA"]
        head=next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        cmap={_norm_text(v):i for i,v in enumerate(head)}
        ciP=cmap.get("PROVINCIA"); ciC=cmap.get("CANTON/CIUDAD"); ciPa=cmap.get("PARROQUIA")
        if ciP is None or ciC is None: return {}
        lastP=lastC=""
        for row in ws.iter_rows(min_row=2, values_only=True):
            P=_norm_text(row[ciP]) or lastP
            C=_norm_text(row[ciC]) or lastC
            Pa=_norm_text(row[ciPa]) if ciPa is not None else ""
            if not P or not C: continue
            lastP, lastC = P, C
            index.setdefault(P, {}).setdefault(C, set())
            if Pa: index[P][C].add(Pa)
        for P in index:
            for C in list(index[P].keys()):
                index[P][C]=sorted(index[P][C]) if index[P][C] else []
        return index
    except: return {}

def cantones_de(prov, idx):
    if not prov: return []
    return sorted(idx.get(_norm_text(prov), {}).keys())

def parroquias_de(prov, cant, idx):
    if not prov or not cant: return []
    return idx.get(_norm_text(prov), {}).get(_norm_text(cant), [])

# ===== DUPLICADOS (solo alerta, no bloquea) =====
def find_col(m,*cands):
    for k,i in m.items():
        for c in cands:
            if c in k: return i
    return None

def buscar_duplicados(cedula_o_ruc, correo, telefono):
    matches=[]
    try: wb=safe_load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except: return matches
    id_new=norm_id(cedula_o_ruc); em_new=norm_email(correo); ph_new=norm_phone(telefono)
    for hoja in ("MECANICO","DISTRIBUIDOR","CONSUMIDOR"):
        if hoja not in wb.sheetnames: continue
        ws=wb[hoja]; hmap=header_map(ws)
        ci=find_col(hmap,"CEDULA","RUC"); ctel=find_col(hmap,"TELEFONO"); ccor=find_col(hmap,"CORREO")
        ccod=find_col(hmap,"CODIGO"); cnom=find_col(hmap,"NOMBRE")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row: continue
            oid=norm_id(str(row[ci])) if ci is not None and row[ci] is not None else ""
            otel=norm_phone(str(row[ctel])) if ctel is not None and row[ctel] is not None else ""
            ocur=norm_email(str(row[ccor])) if ccor is not None and row[ccor] is not None else ""
            dup = (id_new and oid and id_new==oid) or (em_new and ocur and em_new==ocur) or (ph_new and otel and ph_new==otel)
            if dup:
                codigo=str(row[ccod]) if ccod is not None else ""
                nombre=str(row[cnom]) if cnom is not None else ""
                matches.append((hoja, codigo, nombre))
    return matches

# ===== CÓDIGO Y REGISTRO =====
def next_code(prefix):
    ensure_workbook(EXCEL_PATH)
    mx=0
    try:
        wb=safe_load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        if "REGISTRO DE CODIGOS" in wb.sheetnames:
            for (val,) in wb["REGISTRO DE CODIGOS"].iter_rows(min_row=2, max_col=1, values_only=True):
                s=str(val or "").upper(); m=re.search(r"(\d+)$", s) if s.startswith(prefix) else None
                if m: mx=max(mx, int(m.group(1)))
        for hoja,pfx in (("MECANICO","M"),("DISTRIBUIDOR","D"),("CONSUMIDOR","C")):
            if pfx!=prefix or hoja not in wb.sheetnames: continue
            for (val,) in wb[hoja].iter_rows(min_row=2, max_col=1, values_only=True):
                s=str(val or "").upper(); m=re.search(r"(\d+)$", s) if s.startswith(prefix) else None
                if m: mx=max(mx, int(m.group(1)))
    except: pass
    return f"{prefix}{mx+1}"

def upsert_registro_codigo(codigo, ced, nom, tel, tipo, stand):
    ensure_workbook(EXCEL_PATH)
    try: wb=safe_load_workbook(EXCEL_PATH)
    except PermissionError:
        st.info("ℹ️ No se pudo actualizar REGISTRO DE CODIGOS (bloqueado)."); return
    if "REGISTRO DE CODIGOS" not in wb.sheetnames:
        ws=wb.create_sheet("REGISTRO DE CODIGOS"); ws.append(HEADERS["REGISTRO DE CODIGOS"])
    ws=wb["REGISTRO DE CODIGOS"]
    hmap = header_map(ws)
    ci_punt = find_col(hmap, "PUNTAJE")
    ci_ced  = find_col(hmap, "RUC O CEDULA")
    ci_nom  = find_col(hmap, "NOMBRE")
    ci_tel  = find_col(hmap, "TELEFONO")
    ci_tipo = find_col(hmap, "TIPO")
    ci_stand= find_col(hmap, "STAND")
    found=False
    for r in range(2, ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip().upper()==codigo:
            if ci_ced  is not None: ws.cell(r,ci_ced+1).value  = ced
            if ci_nom  is not None: ws.cell(r,ci_nom+1).value  = nom
            if ci_tel  is not None: ws.cell(r,ci_tel+1).value  = tel
            if ci_tipo is not None: ws.cell(r,ci_tipo+1).value = tipo
            if ci_stand is not None: ws.cell(r,ci_stand+1).value= stand
            found=True; break
    if not found:
        row = [""] * (max(hmap.values()) + 1)
        row[0] = codigo
        if ci_punt  is not None: row[ci_punt]  = ""
        if ci_ced   is not None: row[ci_ced]   = ced
        if ci_nom   is not None: row[ci_nom]   = nom
        if ci_tel   is not None: row[ci_tel]   = tel
        if ci_tipo  is not None: row[ci_tipo]  = tipo
        if ci_stand is not None: row[ci_stand] = stand
        ws.append(row)
    try: safe_save_workbook(wb, EXCEL_PATH)
    except PermissionError: st.info("ℹ️ No se pudo guardar (bloqueado).")

def lookup_stand_by_code(wb, code):
    for sheet in ("MECANICO","DISTRIBUIDOR","CONSUMIDOR"):
        if sheet not in wb.sheetnames: continue
        ws = wb[sheet]
        hmap = header_map(ws)
        ci_cod = find_col(hmap, "CODIGO")
        ci_stand = find_col(hmap, "STAND")
        if ci_cod is None or ci_stand is None: continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row: continue
            if str(row[ci_cod] or "").strip().upper() == code:
                return str(row[ci_stand] or "").strip().upper()
    return ""

# ===== CONSULTAS / CARGA DE REGISTROS =====
def _to_int_safe(x, default=0):
    try: return int(x) if x is not None and str(x).strip() != "" else default
    except: return default

@st.cache_data(show_spinner=False)
def load_registros_codigos(path_str: str, mtime: float):
    try:
        wb = safe_load_workbook(Path(path_str), read_only=True, data_only=True)
        if "REGISTRO DE CODIGOS" not in wb.sheetnames: return []
        ws = wb["REGISTRO DE CODIGOS"]
        hmap = header_map(ws)
        ci_cod = find_col(hmap, "CODIGO")
        ci_pun = find_col(hmap, "PUNTAJE")
        ci_ced = find_col(hmap, "RUC O CEDULA")
        ci_nom = find_col(hmap, "NOMBRE")
        ci_tel = find_col(hmap, "TELEFONO")
        ci_tip = find_col(hmap, "TIPO")
        ci_sta = find_col(hmap, "STAND")
        rows=[]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r: continue
            rows.append({
                "CODIGO": str((r[ci_cod] if ci_cod is not None else "") or "").strip(),
                "PUNTAJE": _to_int_safe(r[ci_pun] if ci_pun is not None else 0, 0),
                "RUC O CEDULA": str((r[ci_ced] if ci_ced is not None else "") or "").strip(),
                "NOMBRE": str((r[ci_nom] if ci_nom is not None else "") or "").strip(),
                "TELEFONO": str((r[ci_tel] if ci_tel is not None else "") or "").strip(),
                "TIPO": str((r[ci_tip] if ci_tip is not None else "") or "").strip(),
                "STAND": str((r[ci_sta] if ci_sta is not None else "") or "").strip(),
            })
        return rows
    except: return []

@st.cache_data(show_spinner=False)
def load_registros_premios(path_str: str, mtime: float):
    try:
        wb = safe_load_workbook(Path(path_str), read_only=True, data_only=True)
        if "REGISTRO DE PREMIOS" not in wb.sheetnames: return []
        ws = wb["REGISTRO DE PREMIOS"]
        hmap = header_map(ws)
        ci_cod = find_col(hmap, "CODIGO")
        ci_pre = find_col(hmap, "PREMIO")
        ci_ced = find_col(hmap, "RUC O CEDULA")
        ci_nom = find_col(hmap, "NOMBRE")
        ci_tel = find_col(hmap, "TELEFONO")
        ci_tip = find_col(hmap, "TIPO")
        ci_sta = find_col(hmap, "STAND")
        rows=[]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r: continue
            rows.append({
                "CODIGO": str((r[ci_cod] if ci_cod is not None else "") or "").strip(),
                "PREMIO": str((r[ci_pre] if ci_pre is not None else "") or "").strip(),
                "RUC O CEDULA": str((r[ci_ced] if ci_ced is not None else "") or "").strip(),
                "NOMBRE": str((r[ci_nom] if ci_nom is not None else "") or "").strip(),
                "TELEFONO": str((r[ci_tel] if ci_tel is not None else "") or "").strip(),
                "TIPO": str((r[ci_tip] if ci_tip is not None else "") or "").strip(),
                "STAND": str((r[ci_sta] if ci_sta is not None else "") or "").strip(),
            })
        return rows
    except: return []

def _norm_matchable(s: str) -> str:
    s = (s or "").strip().upper()
    s = (s.replace("Á","A").replace("É","E").replace("Í","I")
           .replace("Ó","O").replace("Ú","U").replace("Ü","U").replace("Ñ","N"))
    return re.sub(r"\s+"," ", s)

def filtrar_por_query(rows: list, q: str, campos=("CODIGO","RUC O CEDULA","NOMBRE","TELEFONO","STAND")):
    if not q: return rows
    qq = _norm_matchable(q)
    out=[]
    for row in rows:
        hit=False
        for c in campos:
            if _norm_matchable(str(row.get(c,""))).find(qq) != -1:
                hit=True; break
        if hit: out.append(row)
    return out

# ===== UI HELPERS =====
def big_code_banner():
    if st.session_state.get("_last_code"):
        st.markdown(
            f"<div style='text-align:center;margin:10px 0 16px 0'>"
            f"<div style='font-size:18px;'>CÓDIGO GENERADO</div>"
            f"<div style='font-size:48px;font-weight:900;color:#e65100'>{st.session_state['_last_code']}</div>"
            f"</div>", unsafe_allow_html=True
        )

def clear_and_rerun(keys):
    for k in keys: st.session_state.pop(k, None)
    if hasattr(st,"rerun"): st.rerun()
    elif hasattr(st,"experimental_rerun"): st.experimental_rerun()

# ===== APP =====
st.title("FORMULARIO DATOS EXPO FERIA")
st.caption(f"Excel (servidor): **{EXCEL_PATH}**")
big_code_banner()

# --- Gestor de Excel en la nube: subir/descargar ---
with st.expander("📁 Excel en el servidor", expanded=False):
    up = st.file_uploader("Cargar/actualizar Excel (.xlsx)", type=["xlsx"], key="excel_up")
    if up is not None:
        with open(EXCEL_PATH, "wb") as f:
            f.write(up.getbuffer())
        st.success(f"Excel cargado/actualizado en: {EXCEL_PATH}")
        if hasattr(st, "rerun"): st.rerun()
    if EXCEL_PATH.exists():
        with open(EXCEL_PATH, "rb") as f:
            st.download_button("⬇️ Descargar Excel actual", f, file_name=EXCEL_PATH.name, key="excel_down")
    else:
        st.info("Aún no hay Excel. Se creará automáticamente al guardar el primer registro.")

prov_index = load_province_index(str(EXCEL_PATH), _mtime(EXCEL_PATH))
provs = sorted(prov_index.keys())

# Tabs (incluye "Consulta")
tabs = st.tabs(["Mecánico","Distribuidor","Consumidor","Puntaje","Premios","Consulta"])

# ---------- MECÁNICO ----------
with tabs[0]:
    st.subheader("Mecánico 🛠️")
    with st.form("m_form"):
        colA,colB = st.columns(2)
        m_nombre = colA.text_input("Nombre y Apellido *", key="m_nombre")
        m_cedula = colB.text_input("RUC o Cédula *", key="m_cedula")
        col1,col2 = st.columns(2)
        m_tel = col1.text_input("Teléfono *", key="m_tel")
        m_correo = col2.text_input("Correo (opcional)", key="m_correo")

        if provs:
            m_prov = st.selectbox("Provincia", [""]+provs, key="m_prov")
            cantones = cantones_de(m_prov, prov_index)
            m_cant = st.selectbox("Cantón / Ciudad", [""]+cantones, key="m_cant") if cantones else st.text_input("Cantón / Ciudad", key="m_cant")
            parroqs = parroquias_de(m_prov, m_cant, prov_index)
            m_parr = st.selectbox("Parroquia", [""]+parroqs, key="m_parr") if parroqs else st.text_input("Parroquia", key="m_parr")
        else:
            m_prov = st.text_input("Provincia", key="m_prov")
            m_cant = st.text_input("Cantón / Ciudad", key="m_cant")
            m_parr = st.text_input("Parroquia", key="m_parr")

        m_dir = st.text_input("Dirección", key="m_dir")
        m_redes = st.text_input("Redes Sociales", key="m_redes")
        col5,col6 = st.columns(2)
        m_dedic = col5.text_input("¿A qué te dedicas?", key="m_dedic")
        m_edad  = col6.number_input("Edad", 0,120,0, key="m_edad")
        m_nom_mec   = st.text_input("Nombre de la mecánica", key="m_nom_mec")
        m_mec_local = st.text_input("Mecánica y local", key="m_mec_local")
        m_visitar   = st.selectbox("¿Quisieras que te visitemos?", ["","SI","NO"], key="m_visitar")
        m_interes   = st.text_area("Productos de interés", key="m_interes")

        m_stand = st.radio("Stand *", ["PANTRO","EXTREMEMAX"], horizontal=True, key="m_stand")

        m_guardar = st.form_submit_button("Guardar MECÁNICO")

    if m_guardar:
        if any(not v for v in [m_nombre, m_cedula, m_tel, m_stand]):
            st.error("Completa: Nombre, Cédula/RUC, Teléfono y Stand."); st.stop()
        if not (validar_cedula_ec(m_cedula) or validar_ruc_natural_ec(m_cedula)):
            st.error("Documento inválido (cédula o RUC natural)."); st.stop()
        if not email_valido(m_correo):
            st.error("Correo inválido."); st.stop()

        dups = buscar_duplicados(m_cedula, m_correo, m_tel)
        if dups:
            lista = "\n".join([f"- {h} | Código {c} | {n}" for h,c,n in dups[:6]])
            st.warning("⚠️ Ya existe un registro con esta Cédula/RUC o Correo o Teléfono:\n" + lista)

        codigo = next_code("M")
        ok = append_row("MECANICO", [
            codigo, m_nombre, m_cedula, m_tel, m_correo,
            m_prov, m_cant, m_parr, m_dir, m_redes, m_dedic,
            int(m_edad) if m_edad else None, m_nom_mec, m_mec_local, m_visitar, m_interes, m_stand
        ])
        if ok:
            upsert_registro_codigo(codigo, m_cedula, m_nombre, m_tel, "MECANICO", m_stand)
            st.session_state["_last_code"]=codigo
            clear_and_rerun([
                "m_nombre","m_cedula","m_tel","m_correo","m_prov","m_cant","m_parr",
                "m_dir","m_redes","m_dedic","m_edad","m_nom_mec","m_mec_local","m_visitar","m_interes","m_stand"
            ])

# ---------- DISTRIBUIDOR ----------
with tabs[1]:
    st.subheader("Distribuidor 🧰")
    with st.form("d_form"):
        colA,colB = st.columns(2)
        d_nombre = colA.text_input("Nombre y Apellido *", key="d_nombre")
        d_cedula = colB.text_input("Cédula o RUC *", key="d_cedula")
        col1,col2 = st.columns(2)
        d_tel = col1.text_input("Teléfono *", key="d_tel")
        d_edad = col2.number_input("Edad", 0,120,0, key="d_edad")

        if provs:
            d_prov = st.selectbox("Provincia", [""]+provs, key="d_prov")
            cantones = cantones_de(d_prov, prov_index)
            d_cant = st.selectbox("Cantón / Ciudad", [""]+cantones, key="d_cant") if cantones else st.text_input("Cantón / Ciudad", key="d_cant")
            parroqs = parroquias_de(d_prov, d_cant, prov_index)
            d_parr = st.selectbox("Parroquia", [""]+parroqs, key="d_parr") if parroqs else st.text_input("Parroquia", key="d_parr")
        else:
            d_prov = st.text_input("Provincia", key="d_prov")
            d_cant = st.text_input("Cantón / Ciudad", key="d_cant")
            d_parr = st.text_input("Parroquia", key="d_parr")

        d_dir = st.text_input("Dirección", key="d_dir")
        d_correo = st.text_input("Correo (opcional)", key="d_correo")
        d_redes = st.text_input("Redes Sociales", key="d_redes")
        d_dedic = st.text_input("¿A qué te dedicas?", key="d_dedic")
        d_rep   = st.text_area("¿Qué repuestos quieres distribuir?", key="d_rep")

        d_stand = st.radio("Stand *", ["PANTRO","EXTREMEMAX"], horizontal=True, key="d_stand")

        d_guardar = st.form_submit_button("Guardar DISTRIBUIDOR")

    if d_guardar:
        if any(not v for v in [d_nombre, d_cedula, d_tel, d_stand]):
            st.error("Completa: Nombre, Cédula/RUC, Teléfono y Stand."); st.stop()
        if not (validar_cedula_ec(d_cedula) or validar_ruc_natural_ec(d_cedula)):
            st.error("Documento inválido (cédula o RUC natural)."); st.stop()
        if not email_valido(d_correo):
            st.error("Correo inválido."); st.stop()

        dups = buscar_duplicados(d_cedula, d_correo, d_tel)
        if dups:
            lista = "\n".join([f"- {h} | Código {c} | {n}" for h,c,n in dups[:6]])
            st.warning("⚠️ Ya existe un registro con esta Cédula/RUC o Correo o Teléfono:\n" + lista)

        codigo = next_code("D")
        ok = append_row("DISTRIBUIDOR", [
            codigo, d_nombre, d_cedula, d_tel, int(d_edad) if d_edad else None,
            d_prov, d_cant, d_parr, d_dir, d_correo, d_redes, d_dedic, d_rep, d_stand
        ])
        if ok:
            upsert_registro_codigo(codigo, d_cedula, d_nombre, d_tel, "DISTRIBUIDOR", d_stand)
            st.session_state["_last_code"]=codigo
            clear_and_rerun([
                "d_nombre","d_cedula","d_tel","d_edad","d_prov","d_cant","d_parr",
                "d_dir","d_correo","d_redes","d_dedic","d_rep","d_stand"
            ])

# ---------- CONSUMIDOR ----------
with tabs[2]:
    st.subheader("Consumidor 🏍️")
    with st.form("c_form"):
        colA,colB = st.columns(2)
        c_nombre = colA.text_input("Nombre y Apellido *", key="c_nombre")
        c_cedula = colB.text_input("Cédula o RUC *", key="c_cedula")
        col1,col2 = st.columns(2)
        c_tel = col1.text_input("Teléfono *", key="c_tel")
        c_edad = col2.number_input("Edad", 0,120,0, key="c_edad")
        c_sexo = st.selectbox("Hombre o Mujer", ["","HOMBRE","MUJER"], key="c_sexo")

        if provs:
            c_prov = st.selectbox("Provincia", [""]+provs, key="c_prov1")
            cantones = cantones_de(c_prov, prov_index)
            c_cant = st.selectbox("Cantón / Ciudad", [""]+cantones, key="c_cant") if cantones else st.text_input("Cantón / Ciudad", key="c_cant")
            parroqs = parroquias_de(c_prov, c_cant, prov_index)
            c_parr = st.selectbox("Parroquia", [""]+parroqs, key="c_parr") if parroqs else st.text_input("Parroquia", key="c_parr")
        else:
            c_prov = st.text_input("Provincia", key="c_prov1")
            c_cant = st.text_input("Cantón / Ciudad", key="c_cant")
            c_parr = st.text_input("Parroquia", key="c_parr")

        c_dir = st.text_input("Dirección", key="c_dir")
        c_dedic = st.text_input("¿A qué te dedicas?", key="c_dedic")
        c_modelo = st.text_input("Modelo de moto que usas", key="c_modelo")
        c_rep = st.text_area("¿Qué repuesto buscas?", key="c_rep")
        c_compra = st.selectbox("¿Has comprado productos ExtremeMax?", ["","SI","NO"], key="c_compra")

        c_stand = st.radio("Stand *", ["PANTRO","EXTREMEMAX"], horizontal=True, key="c_stand")

        c_guardar = st.form_submit_button("Guardar CONSUMIDOR")

    if c_guardar:
        if any(not v for v in [c_nombre, c_cedula, c_tel, c_stand]):
            st.error("Completa: Nombre, Cédula/RUC, Teléfono y Stand."); st.stop()
        if not (validar_cedula_ec(c_cedula) or validar_ruc_natural_ec(c_cedula)):
            st.error("Documento inválido (cédula o RUC natural)."); st.stop()

        dups = buscar_duplicados(c_cedula, "", c_tel)
        if dups:
            lista = "\n".join([f"- {h} | Código {c} | {n}" for h,c,n in dups[:6]])
            st.warning("⚠️ Ya existe un registro con esta Cédula/RUC o Teléfono:\n" + lista)

        codigo = next_code("C")
        ok = append_row("CONSUMIDOR", [
            codigo, c_nombre, c_cedula, c_tel, int(c_edad) if c_edad else None, c_sexo,
            c_prov, c_prov, c_cant, c_parr, c_dir, c_dedic, c_modelo, c_rep, c_compra, c_stand
        ])
        if ok:
            upsert_registro_codigo(codigo, c_cedula, c_nombre, c_tel, "CONSUMIDOR", c_stand)
            st.session_state["_last_code"]=codigo
            clear_and_rerun([
                "c_nombre","c_cedula","c_tel","c_edad","c_sexo","c_prov1","c_cant",
                "c_parr","c_dir","c_dedic","c_modelo","c_rep","c_compra","c_stand"
            ])

# ---------- PUNTAJE ----------
with tabs[3]:
    st.subheader("Asignar puntaje a un código")
    try:
        wb_ro=safe_load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        codes=[]
        if "REGISTRO DE CODIGOS" in wb_ro.sheetnames:
            for row in wb_ro["REGISTRO DE CODIGOS"].iter_rows(min_row=2, values_only=True):
                if row and row[0]: codes.append(str(row[0]).strip())
    except: codes=[]
    col1,col2 = st.columns([2,1])
    cod_sel = col1.selectbox("Código", [""]+sorted(set(codes)), key="puntaje_codigo")
    puntaje = col2.number_input("Puntaje", min_value=0, max_value=100000, step=1, key="puntaje_valor")
    if st.button("Grabar puntaje", key="btn_puntaje"):
        if not cod_sel: st.error("Selecciona un código.")
        else:
            try:
                wb=safe_load_workbook(EXCEL_PATH)
                if "REGISTRO DE CODIGOS" not in wb.sheetnames:
                    ws=wb.create_sheet("REGISTRO DE CODIGOS"); ws.append(HEADERS["REGISTRO DE CODIGOS"])
                ws=wb["REGISTRO DE CODIGOS"]
                hmap = header_map(ws)
                ci_punt = find_col(hmap, "PUNTAJE")
                if ci_punt is None: st.error("No existe la columna PUNTAJE.")
                found=False
                for r in range(2, ws.max_row+1):
                    if str(ws.cell(r,1).value or "").strip()==cod_sel:
                        ws.cell(r,ci_punt+1).value=puntaje; found=True; break
                if not found:
                    def find_data(code):
                        for sheet,tipo in (("MECANICO","MECANICO"),("DISTRIBUIDOR","DISTRIBUIDOR"),("CONSUMIDOR","CONSUMIDOR")):
                            if sheet not in wb.sheetnames: continue
                            ws2=wb[sheet]; h2=header_map(ws2)
                            ci_cod=find_col(h2,"CODIGO")
                            ci_ced=find_col(h2,"CEDULA","RUC")
                            ci_nom=find_col(h2,"NOMBRE")
                            ci_tel=find_col(h2,"TELEFONO")
                            ci_sta=find_col(h2,"STAND")
                            if ci_cod is None: continue
                            for row in ws2.iter_rows(min_row=2, values_only=True):
                                if row and str(row[ci_cod] or "").strip()==code:
                                    return str(row[ci_ced] or ""), str(row[ci_nom] or ""), str(row[ci_tel] or ""), tipo, str(row[ci_sta] or "")
                        return "","","","",""
                    ced, nom, tel, tipo, stand = find_data(cod_sel)
                    hmap = header_map(ws)
                    ci_ced  = find_col(hmap,"RUC O CEDULA")
                    ci_nom  = find_col(hmap,"NOMBRE")
                    ci_tel  = find_col(hmap,"TELEFONO")
                    ci_tip  = find_col(hmap,"TIPO")
                    ci_sta  = find_col(hmap,"STAND")
                    row = [""] * (max(hmap.values()) + 1)
                    row[0] = cod_sel
                    if ci_punt is not None: row[ci_punt] = puntaje
                    if ci_ced  is not None: row[ci_ced]  = ced
                    if ci_nom  is not None: row[ci_nom]  = nom
                    if ci_tel  is not None: row[ci_tel]  = tel
                    if ci_tip  is not None: row[ci_tip]  = tipo
                    if ci_sta  is not None: row[ci_sta]  = stand
                    ws.append(row)
                safe_save_workbook(wb, EXCEL_PATH)
                st.success("✅ Puntaje actualizado.")
            except PermissionError:
                st.error("🔒 No se pudo guardar (archivo bloqueado).")

    st.markdown("---")
    st.subheader("🏁 Top 10 puntajes")
    top = load_registros_codigos(str(EXCEL_PATH), _mtime(EXCEL_PATH))
    if top:
        top = sorted(top, key=lambda x: x.get("PUNTAJE", 0), reverse=True)[:10]
        st.dataframe(top, use_container_width=True, height=360)
    else:
        st.info("Aún no hay registros de puntajes.")

# ---------- PREMIOS ----------
with tabs[4]:
    st.subheader("Registro de premios por código")
    try:
        wb_ro=safe_load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        base={}
        if "REGISTRO DE CODIGOS" in wb_ro.sheetnames:
            ws=wb_ro["REGISTRO DE CODIGOS"]; hmap=header_map(ws)
            ci_cod=find_col(hmap,"CODIGO"); ci_ced=find_col(hmap,"RUC O CEDULA")
            ci_nom=find_col(hmap,"NOMBRE"); ci_tel=find_col(hmap,"TELEFONO")
            ci_tip=find_col(hmap,"TIPO");   ci_sta=find_col(hmap,"STAND")
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row: continue
                cod=str((row[ci_cod] if ci_cod is not None else "") or "").strip()
                if not cod: continue
                base[cod]={
                    "ced": str((row[ci_ced] if ci_ced is not None else "") or ""),
                    "nom": str((row[ci_nom] if ci_nom is not None else "") or ""),
                    "tel": str((row[ci_tel] if ci_tel is not None else "") or ""),
                    "tipo":str((row[ci_tip] if ci_tip is not None else "") or ""),
                    "stand":str((row[ci_sta] if ci_sta is not None else "") or ""),
                }
        if base:
            wb_r = safe_load_workbook(EXCEL_PATH, read_only=True, data_only=True)
            for k,v in base.items():
                if not v.get("stand"):
                    v["stand"] = lookup_stand_by_code(wb_r, k)
        codes=sorted(base.keys())
    except:
        base,codes={},[]
    col1,col2 = st.columns([2,1])
    cod_sel = col1.selectbox("Código", [""]+codes, key="premio_codigo")
    premio  = col2.text_input("Premio *", key="premio_texto")
    if cod_sel:
        d=base.get(cod_sel,{})
        st.write(f"**Nombre:** {d.get('nom','')}")
        st.write(f"**RUC/Cédula:** {d.get('ced','')}")
        st.write(f"**Teléfono:** {d.get('tel','')}")
        st.write(f"**Tipo:** {d.get('tipo','')}")
        st.write(f"**Stand:** {d.get('stand','')}")
    if st.button("Registrar premio", key="btn_premio"):
        if not cod_sel: st.error("Selecciona un código.")
        elif not premio.strip(): st.error("Escribe el premio.")
        else:
            ensure_workbook(EXCEL_PATH)
            try:
                wb=safe_load_workbook(EXCEL_PATH)
                if "REGISTRO DE PREMIOS" not in wb.sheetnames:
                    ws=wb.create_sheet("REGISTRO DE PREMIOS"); ws.append(HEADERS["REGISTRO DE PREMIOS"])
                ws=wb["REGISTRO DE PREMIOS"]
                d=base.get(cod_sel,{"ced":"","nom":"","tel":"","tipo":"","stand":""})
                if not d.get("stand"):
                    d["stand"] = lookup_stand_by_code(wb, cod_sel)
                hmap = header_map(ws)
                ci_cod=find_col(hmap,"CODIGO"); ci_pre=find_col(hmap,"PREMIO")
                ci_ced=find_col(hmap,"RUC O CEDULA"); ci_nom=find_col(hmap,"NOMBRE")
                ci_tel=find_col(hmap,"TELEFONO"); ci_tip=find_col(hmap,"TIPO")
                ci_sta=find_col(hmap,"STAND")
                row = [""] * (max(hmap.values()) + 1)
                if ci_cod is not None: row[ci_cod] = cod_sel
                if ci_pre is not None: row[ci_pre] = premio.strip()
                if ci_ced is not None: row[ci_ced] = d["ced"]
                if ci_nom is not None: row[ci_nom] = d["nom"]
                if ci_tel is not None: row[ci_tel] = d["tel"]
                if ci_tip is not None: row[ci_tip] = d["tipo"]
                if ci_sta is not None: row[ci_sta] = d["stand"]
                ws.append(row)
                safe_save_workbook(wb, EXCEL_PATH)
                st.success("🏆 Premio registrado.")
            except PermissionError:
                st.error("🔒 No se pudo guardar (archivo bloqueado).")

# ---------- CONSULTA ----------
with tabs[5]:
    st.subheader("Consulta rápida 🔎")

    codigos = load_registros_codigos(str(EXCEL_PATH), _mtime(EXCEL_PATH))
    premios = load_registros_premios(str(EXCEL_PATH), _mtime(EXCEL_PATH))

    colA, colB = st.columns([2, 1])
    vista = colB.selectbox("Vista", ["Puntajes", "Premios", "Ambos"])
    q = colA.text_input("Buscar por Código, Cédula/RUC, Nombre, Teléfono o Stand")

    codigos_fil = filtrar_por_query(codigos, q)
    premios_fil = filtrar_por_query(premios, q)

    if vista in ("Puntajes", "Ambos"):
        st.markdown("### Puntajes")
        if codigos_fil:
            if not q:
                codigos_fil = sorted(codigos_fil, key=lambda x: x.get("PUNTAJE", 0), reverse=True)
            st.dataframe(codigos_fil, use_container_width=True, height=360)
        else:
            st.info("Sin resultados de puntajes para la búsqueda.")

    if vista in ("Premios", "Ambos"):
        st.markdown("### Premios")
        if premios_fil:
            st.dataframe(premios_fil, use_container_width=True, height=360)
        else:
            st.info("Sin resultados de premios para la búsqueda.")
